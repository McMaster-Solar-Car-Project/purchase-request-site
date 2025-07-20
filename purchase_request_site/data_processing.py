import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from image_processing import insert_signature_into_worksheet, insert_signature_at_cell
from logging_utils import setup_logger

# Set up logger
logger = setup_logger(__name__)


def copy_expense_report_template(session_folder, user_info, submitted_forms):
    """Copy the expense report template to the session folder and populate with user data"""

    template_path = "excel_templates/expense_report_template.xlsx"

    # Check if template exists
    if not os.path.exists(template_path):
        logger.error(f"Expense report template not found: {template_path}")
        return False

    try:
        # Create destination filename with format: MonthDay-Year-ExpenseReport-FullName
        current_date = datetime.now()
        month_name = current_date.strftime("%B")  # Full month name (e.g., "July")
        day = current_date.strftime("%d").lstrip("0")  # Day without leading zero
        year = current_date.strftime("%Y")  # YYYY format

        # Convert full name to Pascal case (remove spaces, capitalize each word)
        full_name = user_info.get("name", "UnknownUser")
        pascal_name = "".join(word.capitalize() for word in full_name.split())

        output_filename = f"{month_name}{day}-{year}-ExpenseReport-{pascal_name}.xlsx"
        output_path = f"{session_folder}/{output_filename}"

        shutil.copy2(template_path, output_path)
        logger.info(f"Copied expense report template to: {output_path}")

        # Load the copied template and populate with user data
        wb = load_workbook(output_path)

        # Get the active worksheet (assuming first sheet)
        ws = wb.active

        # Populate user information in specified cells
        ws["C2"] = user_info.get("name", "")  # User's name in C2
        ws["F2"] = datetime.now().strftime("%Y-%m-%d")  # Current date in F2
        ws["C3"] = user_info.get("email", "")  # Email in C3
        ws["F3"] = user_info.get("address", "")  # Address in F3

        # Populate expense rows from submitted form data
        try:
            populate_expense_rows_from_submitted_forms(ws, submitted_forms)
        except Exception as e:
            logger.error(f"Failed to populate expense rows from submitted forms: {e}")

        # Insert signature if available
        signature_filename = user_info.get("signature")
        if signature_filename:
            try:
                # Insert signature at cell A19
                insert_signature_at_cell(ws, session_folder, "A19")
                logger.info("Inserted signature into expense report at A19")
            except Exception as e:
                logger.warning(f"Failed to insert signature into expense report: {e}")

        # Save the populated template
        wb.save(output_path)
        wb.close()

        logger.info("Populated expense report with user data:")
        logger.info(f"  - Name: {user_info.get('name', '')}")
        logger.info(f"  - Date: {datetime.now().strftime('%Y-%m-%d')}")
        logger.info(f"  - Email: {user_info.get('email', '')}")
        logger.info(f"  - Address: {user_info.get('address', '')}")

        return True

    except Exception as e:
        logger.error(f"Failed to copy and populate expense report template: {e}")
        return False


def populate_expense_rows_from_submitted_forms(ws, submitted_forms):
    """Populate expense report rows from submitted form data"""

    try:
        current_date = datetime.now().strftime("%Y-%m-%d")
        start_row = 6  # Starting at row 6 as specified
        current_row = start_row

        # Separate Canadian and US forms
        canadian_forms = [
            form for form in submitted_forms if form.get("currency") == "CAD"
        ]
        us_forms = [form for form in submitted_forms if form.get("currency") == "USD"]

        logger.info(
            f"Populating expense report with {len(canadian_forms)} Canadian forms and {len(us_forms)} US forms"
        )

        # Process Canadian forms first
        for i, form in enumerate(canadian_forms):
            row = current_row + i

            # B6, B7, B8... - Date of Receipt
            ws[f"B{row}"] = current_date

            # C6, C7, C8... - Vendor name
            ws[f"C{row}"] = form.get("vendor_name", "")

            # For CAD: D column stays empty, E column stays empty

            # F6, F7, F8... - CDN amount without GST (subtotal)
            subtotal = form.get("subtotal_amount", 0)
            ws[f"F{row}"] = subtotal

            # G6, G7, G8... - CDN amount with GST (total)
            total = form.get("total_amount", 0)
            ws[f"G{row}"] = total

            # H6, H7, H8... - GST/HST amount
            hst_gst = form.get("hst_gst_amount", 0)
            ws[f"H{row}"] = hst_gst

            logger.info(
                f"CAD Row {row}: {form.get('vendor_name', '')} - Subtotal: ${subtotal:.2f}, Total: ${total:.2f}, HST: ${hst_gst:.2f}"
            )

        # Update current row for US forms
        current_row += len(canadian_forms)

        # Process US forms
        for i, form in enumerate(us_forms):
            row = current_row + i

            # B6+, B7+, B8+... - Date of Receipt
            ws[f"B{row}"] = current_date

            # C6+, C7+, C8+... - Vendor name
            ws[f"C{row}"] = form.get("vendor_name", "")

            # D6+, D7+, D8+... - Total foreign amount including taxes (us_total)
            us_total = form.get("us_total", 0)
            ws[f"D{row}"] = us_total

            # E6+, E7+, E8+... - Exchange rate (CAD/USD)
            # Calculate exchange rate from canadian_amount / us_total if both exist
            canadian_amount = form.get("canadian_amount", 0)
            if us_total > 0 and canadian_amount > 0:
                exchange_rate = canadian_amount / us_total
                ws[f"E{row}"] = exchange_rate
            else:
                exchange_rate = 0
                ws[f"E{row}"] = 0
                logger.warning(
                    f"Could not calculate exchange rate for {form.get('vendor_name', '')} - us_total: {us_total}, canadian_amount: {canadian_amount}"
                )

            # F6+, F7+, F8+... - Canadian amount paid (same as G column for US)
            ws[f"F{row}"] = canadian_amount

            # G6+, G7+, G8+... - Canadian amount paid (same as F column for US)
            ws[f"G{row}"] = canadian_amount

            # H6+, H7+, H8+... - HST amount (0 for US purchases)
            ws[f"H{row}"] = 0

            logger.info(
                f"USD Row {row}: {form.get('vendor_name', '')} - USD Total: ${us_total:.2f}, Exchange Rate: {exchange_rate:.4f}, CAD Amount: ${canadian_amount:.2f}"
            )

        total_forms = len(canadian_forms) + len(us_forms)
        logger.info(
            f"Successfully populated {total_forms} total expense rows ({len(canadian_forms)} CAD, {len(us_forms)} USD)"
        )
        return True

    except Exception as e:
        logger.error(f"Error populating expense rows: {e}")
        return False


def create_excel_report(user_info, submitted_forms, session_folder):
    """Create Excel file using the template with multiple tabs for each submitted form"""

    template_path = "excel_templates/purchase_request_template.xlsx"

    # Check if template exists
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Create single output file
    output_filename = "purchase_request.xlsx"
    output_path = f"{session_folder}/{output_filename}"

    # Copy template to session folder
    shutil.copy2(template_path, output_path)

    # Load the copied template
    wb = load_workbook(output_path)

    # Process each submitted form
    for form in submitted_forms:
        # Find the corresponding tab (Receipt1, Receipt2, etc.)
        tab_name = f"Receipt{form['form_number']}"

        # Check if the tab exists
        if tab_name not in wb.sheetnames:
            logger.warning(
                f"Tab '{tab_name}' not found in template, skipping form {form['form_number']}"
            )
            continue

        # Select the appropriate worksheet
        ws = wb[tab_name]

        # Fill in the specified cells
        # B1: Today's date
        ws["B1"] = datetime.now().strftime("%Y-%m-%d")

        # D1: Currency used
        ws["D1"] = form["currency"]

        # B3: Name of submitter
        ws["B3"] = user_info["name"]

        # D3: E-transfer email address
        ws["D3"] = user_info["e_transfer_email"]

        # B4: Team they're on
        ws["B4"] = user_info["team"]

        # B7: Vendor name
        ws["B7"] = form["vendor_name"]

        # B32: Home address
        ws["B32"] = user_info["address"]

        # Populate items starting from row 9 (limit to first 15 items)
        for i, item in enumerate(form["items"][:15]):  # Only process first 15 items
            row = 9 + i  # Start from row 9, increment for each item

            # B9, B10, B11, etc: Item name
            ws[f"B{row}"] = item["name"]

            # C9, C10, C11, etc: Usage/purpose
            ws[f"C{row}"] = item["usage"]

            # D9, D10, D11, etc: Quantity
            ws[f"D{row}"] = item["quantity"]

            # E9, E10, E11, etc: Unit price
            ws[f"E{row}"] = item["unit_price"]

            # F9, F10, F11, etc: Total price
            ws[f"F{row}"] = item["total"]

        # Financial summary (updated row numbers)
        # Note: F24 is NOT the subtotal according to user feedback

        # E25: Tax label (depends on currency)
        if form["currency"] == "USD":
            ws["E25"] = "Taxes"
        else:
            ws["E25"] = "HST/GST"

        # F25: Tax amount (HST/GST for CAD, Taxes for USD)
        ws["F25"] = form["hst_gst_amount"]

        # F26: Shipping
        ws["F26"] = form["shipping_amount"]

        # F27: Total
        ws["F27"] = form["total_amount"]

        # C7 & D7: Conversion Rate (only for USD)
        if form["currency"] == "USD":
            # C7: Conversion Rate label
            ws["C7"] = "Conversion Rate"

            # D7: Calculate conversion rate = Canadian Total / (USD Subtotal + USD Taxes)
            usd_subtotal = sum(
                item["total"] for item in form["items"][:15]
            )  # Only use first 15 items
            usd_taxes = form["hst_gst_amount"]  # This contains USD taxes for USD forms
            usd_total = usd_subtotal + usd_taxes
            canadian_total = form["total_amount"]  # This is the Canadian equivalent

            if usd_total > 0:  # Avoid division by zero
                conversion_rate = canadian_total / usd_total
                ws["D7"] = round(conversion_rate, 4)  # Round to 4 decimal places
            else:
                ws["D7"] = 0

        # Insert signature image
        insert_signature_into_worksheet(ws, user_info, form, session_folder)

    # Save the modified workbook
    wb.save(output_path)
    wb.close()

    # Return information about the single generated file
    return {
        "filename": output_filename,
        "filepath": output_path,
        "forms_processed": len(submitted_forms),
        "tabs_used": [f"Receipt{form['form_number']}" for form in submitted_forms],
    }
