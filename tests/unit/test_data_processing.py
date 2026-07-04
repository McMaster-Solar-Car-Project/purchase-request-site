import re
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

import src.data_processing as data_processing
from src.core.settings import (
    EXCEL_ITEM_END_ROW,
    EXCEL_ITEM_START_ROW,
    EXPENSE_REPORT_END_ROW,
    EXPENSE_REPORT_MIN_ROWS,
    EXPENSE_REPORT_START_ROW,
    MAX_FORMS,
)
from src.models.submissions import Invoice, SubmissionLineItem
from src.models.user_info import SubmissionUserInfo


def _make_form(**overrides) -> Invoice:
    """Build an Invoice with sensible defaults for tests."""
    defaults = {
        "form_number": 1,
        "vendor_name": "Vendor",
        "purchase_date": date(2024, 1, 15),
        "is_usd": False,
        "invoice_filename": "invoice.pdf",
        "invoice_file_location": "/tmp/invoice.pdf",
        "proof_of_payment_filename": None,
        "proof_of_payment_location": None,
        "subtotal_amount": 0.0,
        "discount_amount": 0.0,
        "hst_gst_amount": 0.0,
        "shipping_amount": 0.0,
        "total_cad_amount": 0.0,
        "us_subtotal": 0.0,
        "us_additional_fees": 0.0,
        "items": [
            SubmissionLineItem(name="Item", usage="Test", quantity=1, unit_price=1.0)
        ],
    }
    defaults.update(overrides)
    return Invoice(**defaults)


def _make_user_info() -> SubmissionUserInfo:
    return SubmissionUserInfo(
        name="Test User",
        email="test@example.com",
        e_transfer_email="transfer@example.com",
        address="123 Main St",
        team="Software",
        signature="signature.png",
    )


def _make_items(count: int) -> list[SubmissionLineItem]:
    return [
        SubmissionLineItem(
            name=f"Item {item_number}",
            usage=f"Usage {item_number}",
            quantity=item_number,
            unit_price=1.25,
        )
        for item_number in range(1, count + 1)
    ]


def _expense_report_path(tmp_path: Path) -> Path:
    files = list(tmp_path.glob("*-ExpenseReport-TestUser.xlsx"))
    assert len(files) == 1
    return files[0]


def _as_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    assert isinstance(value, date)
    return value


def test_populate_expense_rows_supports_cad_and_usd() -> None:
    wb = Workbook()
    ws = wb.active

    submitted_forms = [
        _make_form(
            form_number=1,
            vendor_name="CAD Vendor",
            purchase_date=date(2024, 1, 15),
            is_usd=False,
            subtotal_amount=120.0,
            discount_amount=20.0,
            total_cad_amount=113.0,
            hst_gst_amount=13.0,
        ),
        _make_form(
            form_number=2,
            vendor_name="USD Vendor",
            purchase_date=date(2024, 2, 20),
            is_usd=True,
            proof_of_payment_filename="proof.pdf",
            proof_of_payment_location="/tmp/proof.pdf",
            us_subtotal=80.0,
            us_additional_fees=20.0,
            total_cad_amount=135.0,
        ),
    ]

    data_processing.populate_expense_rows_from_submitted_forms(ws, submitted_forms)

    # First row (CAD) starts at row 6.
    assert _as_date(ws["B6"].value) == date(2024, 1, 15)
    assert ws["B6"].number_format == "yyyy-mm-dd"
    assert ws["C6"].value == "CAD Vendor"
    assert ws["F6"].value == 100.0  # subtotal - discount
    assert ws["G6"].value == 113.0
    assert ws["H6"].value == 13.0

    # Second row (USD) is row 7.
    assert _as_date(ws["B7"].value) == date(2024, 2, 20)
    assert ws["B7"].number_format == "yyyy-mm-dd"
    assert ws["C7"].value == "USD Vendor"
    assert ws["D7"].value == 100.0  # US total
    assert ws["E7"].value == 1.35  # Exchange rate
    assert ws["F7"].value == 135.0  # Total amount in CAD
    assert ws["G7"].value == 135.0  # Total amount in CAD
    assert ws["H7"].value == 0  # No HST for US


def test_purchase_request_template_has_twenty_five_receipt_sheets() -> None:
    wb = load_workbook("src/excel_templates/purchase_request_template.xlsx")
    try:
        assert wb.sheetnames == [f"Receipt{i}" for i in range(1, MAX_FORMS + 1)]
        assert wb["Receipt25"]["A58"].value == 50
    finally:
        wb.close()


@pytest.mark.parametrize(
    ("item_count", "expected_hidden_start"),
    [
        (1, 24),
        (15, 24),
        (30, 39),
        (50, None),
    ],
)
def test_create_purchase_request_supports_fifty_item_template_rows(
    monkeypatch, tmp_path, item_count: int, expected_hidden_start: int | None
) -> None:
    signature_calls: list[tuple[str, str, int, int]] = []

    def fake_insert_signature_at_cell(
        ws, _session_folder: str, cell_location: str, width: int, height: int
    ) -> bool:
        signature_calls.append((ws.title, cell_location, width, height))
        return True

    monkeypatch.setattr(
        data_processing, "insert_signature_at_cell", fake_insert_signature_at_cell
    )

    items = _make_items(item_count)
    output_filename = data_processing.create_purchase_request(
        _make_user_info(),
        [
            _make_form(
                items=items,
                subtotal_amount=125.0,
                hst_gst_amount=16.25,
                shipping_amount=10.0,
                total_cad_amount=151.25,
            )
        ],
        str(tmp_path),
    )

    assert re.fullmatch(
        r"[A-Z][a-z]+[1-9][0-9]?-\d{4}-PurchaseRequest-TestUser\.xlsx",
        output_filename,
    )
    assert not (tmp_path / "purchase_request.xlsx").exists()

    wb = load_workbook(tmp_path / output_filename)
    try:
        assert wb.sheetnames == ["Receipt1"]
        ws = wb["Receipt1"]
        assert _as_date(ws["B1"].value) == date(2024, 1, 15)
        assert ws["B1"].number_format == "yyyy-mm-dd"
        assert ws["B67"].value == "123 Main St"
        assert ws["F59"].value == 125.0
        assert ws["F60"].value == 16.25
        assert ws["F61"].value == 10.0
        assert ws["F62"].value == 151.25

        for index, item in enumerate(items, start=EXCEL_ITEM_START_ROW):
            assert ws[f"B{index}"].value == item.name
            assert ws[f"C{index}"].value == item.usage
            assert ws[f"D{index}"].value == item.quantity
            assert ws[f"E{index}"].value == item.unit_price
            assert ws[f"F{index}"].value == item.total

        for row in range(EXCEL_ITEM_START_ROW, EXCEL_ITEM_END_ROW + 1):
            expected_hidden = (
                expected_hidden_start is not None and row >= expected_hidden_start
            )
            assert ws.row_dimensions[row].hidden is expected_hidden
    finally:
        wb.close()

    assert signature_calls == [("Receipt1", "B68", 280, 70)]


def test_create_purchase_request_deletes_unsubmitted_receipt_sheets(
    monkeypatch, tmp_path
) -> None:
    monkeypatch.setattr(
        data_processing,
        "insert_signature_at_cell",
        lambda *_args, **_kwargs: True,
    )

    output_filename = data_processing.create_purchase_request(
        _make_user_info(),
        [
            _make_form(
                form_number=1,
                vendor_name="First Vendor",
                items=_make_items(1),
            ),
            _make_form(
                form_number=3,
                vendor_name="Third Vendor",
                items=_make_items(2),
            ),
        ],
        str(tmp_path),
    )

    wb = load_workbook(tmp_path / output_filename)
    try:
        assert wb.sheetnames == ["Receipt1", "Receipt3"]
        assert wb["Receipt1"]["B7"].value == "First Vendor"
        assert wb["Receipt3"]["B7"].value == "Third Vendor"
    finally:
        wb.close()


def test_create_purchase_request_populates_receipt25_and_removes_unused_sheets(
    monkeypatch, tmp_path
) -> None:
    signature_calls: list[tuple[str, str, int, int]] = []

    def fake_insert_signature_at_cell(
        ws, _session_folder: str, cell_location: str, width: int, height: int
    ) -> bool:
        signature_calls.append((ws.title, cell_location, width, height))
        return True

    monkeypatch.setattr(
        data_processing, "insert_signature_at_cell", fake_insert_signature_at_cell
    )

    output_filename = data_processing.create_purchase_request(
        _make_user_info(),
        [
            _make_form(
                form_number=25,
                vendor_name="Vendor 25",
                subtotal_amount=200.0,
                hst_gst_amount=26.0,
                shipping_amount=5.0,
                total_cad_amount=231.0,
            )
        ],
        str(tmp_path),
    )

    wb = load_workbook(tmp_path / output_filename)
    try:
        assert wb.sheetnames == ["Receipt25"]
        ws = wb["Receipt25"]
        assert _as_date(ws["B1"].value) == date(2024, 1, 15)
        assert ws["B7"].value == "Vendor 25"
        assert ws["B9"].value == "Item"
        assert ws["C9"].value == "Test"
        assert ws["B67"].value == "123 Main St"
        assert ws["F59"].value == 200.0
        assert ws["F60"].value == 26.0
        assert ws["F61"].value == 5.0
        assert ws["F62"].value == 231.0
    finally:
        wb.close()

    assert signature_calls == [("Receipt25", "B68", 280, 70)]


@pytest.mark.parametrize(
    ("form_count", "expected_hidden_start"),
    [
        (1, EXPENSE_REPORT_START_ROW + EXPENSE_REPORT_MIN_ROWS),
        (10, EXPENSE_REPORT_START_ROW + EXPENSE_REPORT_MIN_ROWS),
        (12, EXPENSE_REPORT_START_ROW + 12),
        (25, None),
    ],
)
def test_create_expense_report_supports_twenty_five_rows_and_hides_unused(
    monkeypatch, tmp_path, form_count: int, expected_hidden_start: int | None
) -> None:
    signature_calls: list[tuple[str, int, int]] = []

    def fake_insert_signature_at_cell(
        _ws, _session_folder: str, cell_location: str, width: int, height: int
    ) -> bool:
        signature_calls.append((cell_location, width, height))
        return True

    monkeypatch.setattr(
        data_processing, "insert_signature_at_cell", fake_insert_signature_at_cell
    )

    submitted_forms = [
        _make_form(
            form_number=form_number,
            vendor_name=f"Vendor {form_number}",
            subtotal_amount=100.0 + form_number,
            total_cad_amount=113.0 + form_number,
            hst_gst_amount=13.0,
        )
        for form_number in range(1, form_count + 1)
    ]

    assert data_processing.create_expense_report(
        str(tmp_path), _make_user_info(), submitted_forms
    )

    wb = load_workbook(_expense_report_path(tmp_path), data_only=False)
    try:
        ws = wb.active
        assert _as_date(ws["B6"].value) == date(2024, 1, 15)
        assert ws["B6"].number_format == "yyyy-mm-dd"
        assert ws["C6"].value == "Vendor 1"
        assert ws[f"C{EXPENSE_REPORT_START_ROW + form_count - 1}"].value == (
            f"Vendor {form_count}"
        )
        assert ws["D31"].value == "=SUM(D6:D30)"
        assert ws["E31"].value == "=SUM(E6:E30)"
        assert ws["F31"].value == "=SUM(F6:F30)"
        assert ws["G31"].value == "=SUM(G6:G30)"
        assert ws["H31"].value == "=SUM(H6:H30)"
        assert ws["G32"].value == "=G31"

        for row in range(EXPENSE_REPORT_START_ROW, EXPENSE_REPORT_END_ROW + 1):
            expected_hidden = (
                expected_hidden_start is not None and row >= expected_hidden_start
            )
            assert ws.row_dimensions[row].hidden is expected_hidden
    finally:
        wb.close()

    assert signature_calls == [("A34", 200, 60)]
