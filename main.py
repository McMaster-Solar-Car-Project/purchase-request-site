import os
import time
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, Request, Form, File, UploadFile
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from dotenv import load_dotenv
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan manager - runs on startup and shutdown"""
    # Startup: Create required directories
    required_dirs = ["sessions", "static", "templates"]
    
    for directory in required_dirs:
        os.makedirs(directory, exist_ok=True)
        print(f"✓ Ensured directory exists: {directory}/")
    
    print("🚀 Application startup complete - all directories ready!")
    
    yield  # Application runs here
    
    # Shutdown: Cleanup code would go here if needed
    print("👋 Application shutting down...")

app = FastAPI(title="Purchase Request Site", lifespan=lifespan)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/sessions", StaticFiles(directory="sessions"), name="sessions")

# Set up templates
templates = Jinja2Templates(directory="templates")

@app.get("/")
async def home(request: Request, success: str = None):
    success_message = None
    if success:
        success_message = "All your purchase requests have been submitted successfully! We'll be in touch soon."
    
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "title": "Purchase Request Site",
            "heading": "Enter your purchase request information",
            "success_message": success_message,
            "google_api_key": os.getenv("GOOGLE_PLACES_API_KEY"),
        },
    )

def create_session_folder(name):
    """Create a unique session folder name with timestamp and user name"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = name.replace(' ', '_').lower()
    session_folder = f"sessions/{timestamp}_{safe_name}"
    
    # Create the session directory if it doesn't exist
    os.makedirs(session_folder, exist_ok=True)
    
    return session_folder

@app.post("/submit-request")
async def submit_request(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    e_transfer_email: str = Form(...),
    address: str = Form(...),
    team: str = Form(...),
    signature: UploadFile = File(...),
):
    # Create session folder for this user
    session_folder = create_session_folder(name)
    
    # Save signature file in session folder
    signature_extension = signature.filename.split('.')[-1] if '.' in signature.filename else 'png'
    signature_filename = f"signature.{signature_extension}"
    signature_location = f"{session_folder}/{signature_filename}"
    
    # Save the signature file
    with open(signature_location, "wb") as file_object:
        content = await signature.read()
        file_object.write(content)
    
    # Redirect to dashboard with user information including session folder
    return RedirectResponse(
        url=f"/dashboard?name={name}&email={email}&e_transfer_email={e_transfer_email}&address={address}&team={team}&session_folder={session_folder}&signature={signature_filename}",
        status_code=303
    )

@app.get("/dashboard")
async def dashboard(
    request: Request,
    name: str,
    email: str,
    e_transfer_email: str,
    address: str,
    team: str,
    session_folder: str,
    signature: str,
):
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "title": "Purchase Request Site",
            "name": name,
            "email": email,
            "e_transfer_email": e_transfer_email,
            "address": address,
            "team": team,
            "session_folder": session_folder,
            "signature": signature,
        },
    )

def create_excel_export(user_info, submitted_forms, session_folder):
    """Create an Excel file with all submitted form data in the session folder"""
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # User Information Section
    ws_summary['A1'] = "PURCHASE REQUEST SUBMISSION SUMMARY"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:G1')
    
    ws_summary['A3'] = "Submission Date:"
    ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary['A4'] = "Submitter Name:"
    ws_summary['B4'] = user_info['name']
    
    ws_summary['A5'] = "McMaster Email:"
    ws_summary['B5'] = user_info['email']
    
    ws_summary['A6'] = "E-Transfer Email:"
    ws_summary['B6'] = user_info['e_transfer_email']
    
    ws_summary['A7'] = "Address:"
    ws_summary['B7'] = user_info['address']
    
    ws_summary['A8'] = "Team/Department:"
    ws_summary['B8'] = user_info['team']
    
    ws_summary['A9'] = "Digital Signature:"
    ws_summary['B9'] = user_info['signature']
    
    # Forms Summary Section
    ws_summary['A11'] = "FORMS SUBMITTED"
    ws_summary['A11'].font = Font(bold=True, size=14)
    
    # Summary headers
    summary_headers = ["Form #", "Vendor", "Currency", "Total Amount", "# of Items", "Invoice File", "Proof of Payment"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=13, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Summary data
    total_overall = 0
    for i, form in enumerate(submitted_forms, 14):
        ws_summary.cell(row=i, column=1, value=form['form_number'])
        ws_summary.cell(row=i, column=2, value=form['vendor_name'])
        ws_summary.cell(row=i, column=3, value=form['currency'])
        ws_summary.cell(row=i, column=4, value=f"${form['total_amount']:.2f} {form['currency']}")
        ws_summary.cell(row=i, column=5, value=len(form['items']))
        ws_summary.cell(row=i, column=6, value=form['invoice_filename'])
        ws_summary.cell(row=i, column=7, value=form['proof_of_payment_filename'] or "N/A")
        total_overall += form['total_amount']
    
    # Total row
    total_row = 14 + len(submitted_forms)
    ws_summary.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=3, value=f"${total_overall:.2f}").font = Font(bold=True)
    
    # Create detailed sheet for each form
    for form in submitted_forms:
        ws_detail = wb.create_sheet(title=f"Form {form['form_number']}")
        
        # Form header
        ws_detail['A1'] = f"PURCHASE REQUEST #{form['form_number']}"
        ws_detail['A1'].font = Font(bold=True, size=16)
        ws_detail.merge_cells('A1:F1')
        
        # Vendor information
        ws_detail['A3'] = "Vendor/Store:"
        ws_detail['B3'] = form['vendor_name']
        
        ws_detail['A4'] = "Currency:"
        ws_detail['B4'] = form['currency']
        
        ws_detail['A5'] = "Invoice File:"
        ws_detail['B5'] = form['invoice_filename']
        
        # Add proof of payment info if exists
        row_offset = 0
        if form['proof_of_payment_filename']:
            ws_detail['A6'] = "Proof of Payment:"
            ws_detail['B6'] = form['proof_of_payment_filename']
            row_offset = 1
        
        # Financial breakdown (adjust row numbers based on proof of payment)
        ws_detail[f'A{7 + row_offset}'] = "FINANCIAL BREAKDOWN"
        ws_detail[f'A{7 + row_offset}'].font = Font(bold=True, size=14)
        
        if form['currency'] == "USD":
            # USD breakdown
            ws_detail[f'A{8 + row_offset}'] = "US Total:"
            ws_detail[f'B{8 + row_offset}'] = f"${form['us_total']:.2f} USD"
            
            ws_detail[f'A{9 + row_offset}'] = "Canadian Amount:"
            ws_detail[f'B{9 + row_offset}'] = f"${form['canadian_amount']:.2f} CAD"
            
            ws_detail[f'A{10 + row_offset}'] = "REIMBURSEMENT TOTAL:"
            ws_detail[f'A{10 + row_offset}'].font = Font(bold=True)
            ws_detail[f'B{10 + row_offset}'] = f"${form['canadian_amount']:.2f} CAD"
            ws_detail[f'B{10 + row_offset}'].font = Font(bold=True)
            
            # Adjust items section row
            items_row = 12 + row_offset
        else:
            # CAD breakdown
            ws_detail[f'A{8 + row_offset}'] = "Subtotal:"
            ws_detail[f'B{8 + row_offset}'] = f"${form['subtotal_amount']:.2f} {form['currency']}"
            
            ws_detail[f'A{9 + row_offset}'] = "Discount:"
            ws_detail[f'B{9 + row_offset}'] = f"-${form['discount_amount']:.2f} {form['currency']}"
            
            # Use appropriate tax label based on currency
            tax_label = "Taxes" if form['currency'] == "USD" else "HST/GST"
            ws_detail[f'A{10 + row_offset}'] = f"{tax_label}:"
            ws_detail[f'B{10 + row_offset}'] = f"${form['hst_gst_amount']:.2f} {form['currency']}"
            
            ws_detail[f'A{11 + row_offset}'] = "Shipping:"
            ws_detail[f'B{11 + row_offset}'] = f"${form['shipping_amount']:.2f} {form['currency']}"
            
            ws_detail[f'A{12 + row_offset}'] = "TOTAL:"
            ws_detail[f'A{12 + row_offset}'].font = Font(bold=True)
            ws_detail[f'B{12 + row_offset}'] = f"${form['total_amount']:.2f} {form['currency']}"
            ws_detail[f'B{12 + row_offset}'].font = Font(bold=True)
            
            # Adjust items section row
            items_row = 14 + row_offset
        
        # Items section
        ws_detail[f'A{items_row}'] = "ITEMS PURCHASED"
        ws_detail[f'A{items_row}'].font = Font(bold=True, size=14)
        
        # Item headers
        item_headers = ["Item #", "Item Name", "Usage/Purpose", "Quantity", f"Unit Price ({form['currency']})", f"Total ({form['currency']})"]
        for col, header in enumerate(item_headers, 1):
            cell = ws_detail.cell(row=items_row + 2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Item data
        for i, item in enumerate(form['items'], items_row + 3):
            ws_detail.cell(row=i, column=1, value=i - (items_row + 2))
            ws_detail.cell(row=i, column=2, value=item['name'])
            ws_detail.cell(row=i, column=3, value=item['usage'])
            ws_detail.cell(row=i, column=4, value=item['quantity'])
            ws_detail.cell(row=i, column=5, value=f"${item['unit_price']:.2f}")
            ws_detail.cell(row=i, column=6, value=f"${item['total']:.2f}")
        
        # Auto-adjust column widths
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            ws_detail.column_dimensions[column_letter].width = 20
    
    # Auto-adjust column widths for summary
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws_summary.column_dimensions[column_letter].width = 20
    
    # Save Excel file in session folder
    filename = "purchase_requests.xlsx"
    filepath = f"{session_folder}/{filename}"
    
    # Save the workbook
    wb.save(filepath)
    
    return filename, filepath

@app.post("/submit-all-requests")
async def submit_all_requests(request: Request):
    # Get form data
    form_data = await request.form()
    
    # Extract user information
    name = form_data.get("name")
    email = form_data.get("email")
    e_transfer_email = form_data.get("e_transfer_email")
    address = form_data.get("address")
    team = form_data.get("team")
    session_folder = form_data.get("session_folder")
    signature_filename = form_data.get("signature")
    
    submitted_forms = []
    
    # Process each of the 10 possible forms
    for form_num in range(1, 11):
        vendor_name = form_data.get(f"vendor_name_{form_num}")
        invoice_file = form_data.get(f"invoice_file_{form_num}")
        proof_of_payment_file = form_data.get(f"proof_of_payment_{form_num}")
        currency = form_data.get(f"currency_{form_num}", "CAD")  # Default to CAD if not specified
        
        # Skip empty forms (no vendor name or invoice)
        if not vendor_name or not invoice_file or not hasattr(invoice_file, 'filename'):
            continue
        
        # For USD purchases, proof of payment is required
        if currency == "USD" and (not proof_of_payment_file or not hasattr(proof_of_payment_file, 'filename')):
            print(f"Warning: Form {form_num} in USD currency missing proof of payment - skipping")
            continue
            
        # Extract financial data based on currency
        if currency == "USD":
            # For USD, use simplified breakdown
            us_total = float(form_data.get(f"us_total_{form_num}") or 0)
            canadian_amount = float(form_data.get(f"canadian_amount_{form_num}") or 0)
            
            # Set other fields to 0 for USD
            subtotal_amount = 0
            discount_amount = 0
            hst_gst_amount = 0
            shipping_amount = 0
            total_amount = canadian_amount  # Use Canadian amount as total for reimbursement
        else:
            # For CAD, use detailed breakdown
            subtotal_amount = float(form_data.get(f"subtotal_amount_{form_num}") or 0)
            discount_amount = float(form_data.get(f"discount_amount_{form_num}") or 0)
            hst_gst_amount = float(form_data.get(f"hst_gst_amount_{form_num}") or 0)
            shipping_amount = float(form_data.get(f"shipping_amount_{form_num}") or 0)
            total_amount = float(form_data.get(f"total_amount_{form_num}") or 0)
            
            # Set USD fields to 0 for CAD
            us_total = 0
            canadian_amount = 0
        
        # Extract items for this form
        items = []
        item_num = 1
        while True:
            item_name = form_data.get(f"item_name_{form_num}_{item_num}")
            if not item_name:
                break
                
            item_usage = form_data.get(f"item_usage_{form_num}_{item_num}")
            item_quantity = form_data.get(f"item_quantity_{form_num}_{item_num}")
            item_price = form_data.get(f"item_price_{form_num}_{item_num}")
            item_total = form_data.get(f"item_total_{form_num}_{item_num}")
            
            if item_name and item_usage and item_quantity and item_price:
                items.append({
                    "name": item_name,
                    "usage": item_usage,
                    "quantity": int(item_quantity),
                    "unit_price": float(item_price),
                    "total": float(item_total) if item_total else 0
                })
            
            item_num += 1
        
        # Skip forms with no items
        if not items:
            continue
            
        # Save uploaded invoice file in session folder
        invoice_extension = invoice_file.filename.split('.')[-1] if '.' in invoice_file.filename else 'pdf'
        invoice_filename = f"{form_num}_invoice.{invoice_extension}"
        invoice_file_location = f"{session_folder}/{invoice_filename}"
        
        # Save the invoice file
        with open(invoice_file_location, "wb") as file_object:
            content = await invoice_file.read()
            file_object.write(content)
        
        # Save proof of payment file if provided (required for USD)
        proof_of_payment_filename = None
        proof_of_payment_location = None
        if proof_of_payment_file and hasattr(proof_of_payment_file, 'filename'):
            payment_extension = proof_of_payment_file.filename.split('.')[-1] if '.' in proof_of_payment_file.filename else 'pdf'
            proof_of_payment_filename = f"{form_num}_proof_of_payment.{payment_extension}"
            proof_of_payment_location = f"{session_folder}/{proof_of_payment_filename}"
            
            # Save the proof of payment file
            with open(proof_of_payment_location, "wb") as file_object:
                content = await proof_of_payment_file.read()
                file_object.write(content)
        
        # Store form data
        form_submission = {
            "form_number": form_num,
            "vendor_name": vendor_name,
            "currency": currency,
            "invoice_filename": invoice_filename,
            "invoice_file_location": invoice_file_location,
            "proof_of_payment_filename": proof_of_payment_filename,
            "proof_of_payment_location": proof_of_payment_location,
            "subtotal_amount": subtotal_amount,
            "discount_amount": discount_amount,
            "hst_gst_amount": hst_gst_amount,
            "shipping_amount": shipping_amount,
            "total_amount": total_amount,
            "us_total": us_total,
            "canadian_amount": canadian_amount,
            "items": items
        }
        
        submitted_forms.append(form_submission)
    
    # Print all submitted forms
    if submitted_forms:
        # Create Excel export in session folder
        user_info = {
            'name': name,
            'email': email,
            'e_transfer_email': e_transfer_email,
            'address': address,
            'team': team,
            'signature': signature_filename
        }
        
        excel_filename, excel_filepath = create_excel_export(user_info, submitted_forms, session_folder)
        
        print(f"Bulk submission received from:")
        print(f"Name: {name}")
        print(f"McMaster Email: {email}")
        print(f"E-Transfer Email: {e_transfer_email}")
        print(f"Address: {address}")
        print(f"Team: {team}")
        print(f"Session Folder: {session_folder}")
        print(f"Digital Signature: {signature_filename} (saved to {session_folder}/{signature_filename})")
        print(f"Excel Export: {excel_filename} (saved to {excel_filepath})")
        print(f"")
        print(f"Number of forms submitted: {len(submitted_forms)}")
        print("=" * 60)
        
        for form in submitted_forms:
            print(f"Purchase Request #{form['form_number']}:")
            print(f"  Vendor: {form['vendor_name']}")
            print(f"  Currency: {form['currency']}")
            print(f"  Invoice File: {form['invoice_filename']} (saved to {form['invoice_file_location']})")
            if form['proof_of_payment_filename']:
                print(f"  Proof of Payment: {form['proof_of_payment_filename']} (saved to {form['proof_of_payment_location']})")
            print(f"  Financial Breakdown:")
            
            if form['currency'] == "USD":
                # Show USD breakdown
                print(f"    US Total: ${form['us_total']:.2f} USD")
                print(f"    Canadian Amount: ${form['canadian_amount']:.2f} CAD")
                print(f"    Reimbursement Total: ${form['canadian_amount']:.2f} CAD")
            else:
                # Show CAD breakdown
                print(f"    Subtotal: ${form['subtotal_amount']:.2f} {form['currency']}")
                print(f"    Discount: -${form['discount_amount']:.2f} {form['currency']}")
                
                # Use appropriate tax label based on currency
                tax_label = "Taxes" if form['currency'] == "USD" else "HST/GST"
                print(f"    {tax_label}: ${form['hst_gst_amount']:.2f} {form['currency']}")
                
                print(f"    Shipping: ${form['shipping_amount']:.2f} {form['currency']}")
                print(f"    Total Amount: ${form['total_amount']:.2f} {form['currency']}")
            
            print(f"  Items:")
            for i, item in enumerate(form['items'], 1):
                print(f"    {i}. {item['name']}")
                print(f"       Usage: {item['usage']}")
                print(f"       Quantity: {item['quantity']}")
                print(f"       Unit Price: ${item['unit_price']:.2f} {form['currency']}")
                print(f"       Total: ${item['total']:.2f} {form['currency']}")
            print("-" * 40)
        
        print("=" * 60)
    else:
        print("No forms were submitted (all forms were empty)")
    
    # Redirect back to home with success message
    return RedirectResponse(url="/?success=true", status_code=303)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app,
        host=os.getenv("HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", 8000)),
        reload=os.getenv("DEBUG", "false").lower() == "true",
    )
