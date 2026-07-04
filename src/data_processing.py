import shutil
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.logging_utils import setup_logger
from src.core.settings import (
    EXCEL_ITEM_END_ROW,
    EXCEL_ITEM_START_ROW,
    EXPENSE_REPORT_END_ROW,
    EXPENSE_REPORT_MIN_ROWS,
    EXPENSE_REPORT_START_ROW,
    MIN_EXCEL_ITEM_ROWS,
)
from src.image_processing import insert_signature_at_cell
from src.models.submissions import Invoice
from src.models.user_info import SubmissionUserInfo

logger = setup_logger(__name__)


@contextmanager
def _copied_template_workbook(
    template_path: str | Path, output_path: str | Path
) -> Iterator[Workbook]:
    """Yield a workbook copied from a template, cleaning up partial output on failure."""
    wb = None
    try:
        shutil.copy2(template_path, output_path)
        wb = load_workbook(output_path)
        yield wb
        wb.save(output_path)
    except Exception:
        try:
            Path(output_path).unlink(missing_ok=True)
        except OSError as e:
            logger.warning(f"Failed to remove partial output {output_path}: {e}")
        raise
    finally:
        if wb is not None:
            wb.close()


def _hide_unused_rows(
    ws: Worksheet,
    used_count: int,
    *,
    start_row: int,
    end_row: int,
    min_visible_rows: int,
    overflow_message: str,
) -> None:
    available_rows = end_row - start_row + 1
    if used_count > available_rows:
        raise ValueError(overflow_message)

    first_hidden_row = start_row + max(min_visible_rows, used_count)
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].hidden = row >= first_hidden_row


def create_expense_report(
    session_folder: str,
    user_info: SubmissionUserInfo,
    submitted_forms: list[Invoice],
) -> bool:
    """Copy the expense report template to the session folder and populate with user data."""
    template_path = "src/excel_templates/expense_report_template.xlsx"
    if not Path(template_path).exists():
        logger.error(f"Expense report template not found: {template_path}")
        return False

    now = datetime.now()
    day = now.strftime("%d").lstrip("0")
    pascal_name = "".join(word.capitalize() for word in user_info.name.split())
    output_filename = f"{now.strftime('%B')}{day}-{now.strftime('%Y')}-ExpenseReport-{pascal_name}.xlsx"
    output_path = f"{session_folder}/{output_filename}"

    try:
        with _copied_template_workbook(template_path, output_path) as wb:
            ws = wb.active
            ws["C2"] = user_info.name
            ws["F2"] = now.strftime("%Y-%m-%d")
            ws["C3"] = user_info.email
            ws["F3"] = user_info.address

            populate_expense_rows_from_submitted_forms(ws, submitted_forms)
            _hide_unused_rows(
                ws,
                len(submitted_forms),
                start_row=EXPENSE_REPORT_START_ROW,
                end_row=EXPENSE_REPORT_END_ROW,
                min_visible_rows=EXPENSE_REPORT_MIN_ROWS,
                overflow_message=(
                    f"Expense report has {len(submitted_forms)} invoices, "
                    "but the template does not have enough rows"
                ),
            )

            try:
                insert_signature_at_cell(ws, session_folder, "A34", 200, 60)
            except Exception as e:
                logger.warning(f"Failed to insert signature into expense report: {e}")

        return True
    except Exception:
        logger.exception("Failed to create expense report")
        return False


def populate_expense_rows_from_submitted_forms(
    ws: Worksheet, submitted_forms: list[Invoice]
) -> None:
    """Populate expense report rows from submitted form data."""
    for i, form in enumerate(submitted_forms):
        row = EXPENSE_REPORT_START_ROW + i
        ws[f"B{row}"] = form.purchase_date
        ws[f"B{row}"].number_format = "yyyy-mm-dd"
        ws[f"C{row}"] = form.vendor_name

        if not form.is_usd:
            ws[f"F{row}"] = form.subtotal_amount - form.discount_amount
            ws[f"G{row}"] = form.total_cad_amount
            ws[f"H{row}"] = form.hst_gst_amount
        else:
            ws[f"D{row}"] = form.us_total
            ws[f"E{row}"] = form.exchange_rate
            ws[f"F{row}"] = form.total_cad_amount
            ws[f"G{row}"] = form.total_cad_amount
            ws[f"H{row}"] = 0


def _delete_unused_receipt_sheets(wb, submitted_forms: list[Invoice]) -> None:
    submitted_sheet_names = {f"Receipt{form.form_number}" for form in submitted_forms}

    for ws in list(wb.worksheets):
        if ws.title not in submitted_sheet_names:
            wb.remove(ws)


def create_purchase_request(
    user_info: SubmissionUserInfo,
    submitted_forms: list[Invoice],
    session_folder: str,
) -> str:
    """Create Purchase Request using a template with one tab per submitted form."""
    template_path = "src/excel_templates/purchase_request_template.xlsx"
    if not Path(template_path).exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    now = datetime.now()
    day = now.strftime("%d").lstrip("0")
    pascal_name = "".join(word.capitalize() for word in user_info.name.split())
    output_filename = f"{now.strftime('%B')}{day}-{now.strftime('%Y')}-PurchaseRequest-{pascal_name}.xlsx"
    output_path = f"{session_folder}/{output_filename}"
    with _copied_template_workbook(template_path, output_path) as wb:
        _delete_unused_receipt_sheets(wb, submitted_forms)

        for ws in wb.worksheets:
            _hide_unused_rows(
                ws,
                0,
                start_row=EXCEL_ITEM_START_ROW,
                end_row=EXCEL_ITEM_END_ROW,
                min_visible_rows=MIN_EXCEL_ITEM_ROWS,
                overflow_message="Purchase request template does not have enough rows",
            )

        for form in submitted_forms:
            tab_name = f"Receipt{form.form_number}"

            if tab_name not in wb.sheetnames:
                logger.warning(
                    f"Tab '{tab_name}' not found in template, skipping form {form.form_number}"
                )
                continue

            ws = wb[tab_name]
            _hide_unused_rows(
                ws,
                len(form.items),
                start_row=EXCEL_ITEM_START_ROW,
                end_row=EXCEL_ITEM_END_ROW,
                min_visible_rows=MIN_EXCEL_ITEM_ROWS,
                overflow_message=(
                    f"Form {form.form_number} has {len(form.items)} items, "
                    f"but the template supports at most "
                    f"{EXCEL_ITEM_END_ROW - EXCEL_ITEM_START_ROW + 1}"
                ),
            )

            ws["B1"] = form.purchase_date
            ws["B1"].number_format = "yyyy-mm-dd"
            ws["D1"] = "USD" if form.is_usd else "CAD"
            ws["B3"] = user_info.name
            ws["D3"] = user_info.e_transfer_email
            ws["B4"] = user_info.team
            ws["B7"] = form.vendor_name
            ws["B67"] = user_info.address

            for i, item in enumerate(form.items):
                row = EXCEL_ITEM_START_ROW + i
                ws[f"B{row}"] = item.name
                ws[f"C{row}"] = item.usage
                ws[f"D{row}"] = item.quantity
                ws[f"E{row}"] = item.unit_price
                ws[f"F{row}"] = item.total

            ws["F59"] = form.us_subtotal if form.is_usd else form.subtotal_amount
            ws["F60"] = form.us_additional_fees if form.is_usd else form.hst_gst_amount
            ws["F61"] = form.us_total if form.is_usd else form.shipping_amount
            ws["F62"] = form.total_cad_amount

            if form.is_usd:
                ws["D7"] = round(form.exchange_rate, 4)

            insert_signature_at_cell(ws, session_folder, "B68", 280, 70)

    return output_filename
