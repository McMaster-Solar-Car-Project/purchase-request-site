import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from image_processing import insert_signature_into_worksheet


def create_excel_report(user_info, submitted_forms, session_folder):
    """Create Excel file using the template with multiple tabs for each submitted form"""

    template_path = "excel_templates/purchase_request_template.xlsx"

    # Check if template exists
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")

    # Create single output file
    output_filename = (
        f"purchase_request.xlsx"
    )
    output_path = f"{session_folder}/{output_filename}"

    # Copy template to session folder
    shutil.copy2(template_path, output_path)

    # Load the copied template
    wb = load_workbook(output_path)

    # Process each submitted form
    for form in submitted_forms:
        # Find the corresponding tab (Receipt1, Receipt2, etc.)
        tab_name = f"Receipt{form['form_number']}"

        # Check if the tab exists
        if tab_name not in wb.sheetnames:
            print(
                f"Warning: Tab '{tab_name}' not found in template, skipping form {form['form_number']}"
            )
            continue

        # Select the appropriate worksheet
        ws = wb[tab_name]

        # Fill in the specified cells
        # B1: Today's date
        ws["B1"] = datetime.now().strftime("%Y-%m-%d")

        # D1: Currency used
        ws["D1"] = form["currency"]

        # B3: Name of submitter
        ws["B3"] = user_info["name"]

        # D3: E-transfer email address
        ws["D3"] = user_info["e_transfer_email"]

        # B4: Team they're on
        ws["B4"] = user_info["team"]

        # B7: Vendor name
        ws["B7"] = form["vendor_name"]

        # B24: Home address
        ws["B24"] = user_info["address"]

        # Populate items starting from row 9
        for i, item in enumerate(form["items"]):
            row = 9 + i  # Start from row 9, increment for each item

            # B9, B10, B11, etc: Item name
            ws[f"B{row}"] = item["name"]

            # C9, C10, C11, etc: Usage/purpose
            ws[f"C{row}"] = item["usage"]

            # D9, D10, D11, etc: Quantity
            ws[f"D{row}"] = item["quantity"]

            # E9, E10, E11, etc: Unit price
            ws[f"E{row}"] = item["unit_price"]

            # F9, F10, F11, etc: Total price
            ws[f"F{row}"] = item["total"]

        # Financial summary
        # F16: Subtotal (different logic for USD vs CAD)
        if form["currency"] == "USD":
            # For USD: Calculate subtotal from item totals (excluding taxes)
            usd_subtotal = sum(item["total"] for item in form["items"])
            ws["F16"] = usd_subtotal
        else:
            # For CAD: Use the calculated subtotal amount
            ws["F16"] = form["subtotal_amount"]

        # E17: Tax label (depends on currency)
        if form["currency"] == "USD":
            ws["E17"] = "Taxes"
        else:
            ws["E17"] = "HST/GST"

        # F17: Tax amount (HST/GST for CAD, Taxes for USD)
        ws["F17"] = form["hst_gst_amount"]

        # F18: Shipping
        ws["F18"] = form["shipping_amount"]

        # F19: Total
        ws["F19"] = form["total_amount"]

        # C7 & D7: Conversion Rate (only for USD)
        if form["currency"] == "USD":
            # C7: Conversion Rate label
            ws["C7"] = "Conversion Rate"

            # D7: Calculate conversion rate = Canadian Total / (USD Subtotal + USD Taxes)
            usd_subtotal = sum(item["total"] for item in form["items"])
            usd_taxes = form["hst_gst_amount"]  # This contains USD taxes for USD forms
            usd_total = usd_subtotal + usd_taxes
            canadian_total = form["total_amount"]  # This is the Canadian equivalent

            if usd_total > 0:  # Avoid division by zero
                conversion_rate = canadian_total / usd_total
                ws["D7"] = round(conversion_rate, 4)  # Round to 4 decimal places
            else:
                ws["D7"] = 0

        # Insert signature image
        insert_signature_into_worksheet(ws, user_info, form, session_folder)

    # Save the modified workbook
    wb.save(output_path)
    wb.close()

    # Return information about the single generated file
    return {
        "filename": output_filename,
        "filepath": output_path,
        "forms_processed": len(submitted_forms),
        "tabs_used": [f"Receipt{form['form_number']}" for form in submitted_forms],
    }
